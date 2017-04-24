from django.core.management import BaseCommand

from planning.models import *
from openpyxl import load_workbook


def import_PlannedActivities(excel_file, year):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    worksheet_name = "AWP 2017"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:AC%s' % (workbook_results.min_row + 4, workbook_results.max_row)):
            pa = PlannedActivities()
            pa.area = row[0].value
            pa.description = row[2].value
            if row[4].value:
                pa.fund = row[4].value
            else:
                pa.fund = False
            pa.priority = row[5].value
            pa.qtr1 = row[25].value
            pa.qtr2 = row[26].value
            pa.qtr3 = row[27].value
            pa.qtr4 = row[28].value
            pa.year = year
            pa.save()

class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """ Import PlannedActivities """

    def handle(self, *args, **options):
        excel_file = args[0]
        year = args[1]
        import_PlannedActivities(excel_file, year)