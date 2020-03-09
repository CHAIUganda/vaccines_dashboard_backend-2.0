from datetime import date
from django.core.management import BaseCommand
from django.db import IntegrityError

from cold_chain.models import *
from openpyxl import load_workbook


def import_facilities(excel_file):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    worksheet_name = "tbl_facilities"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:F%s' % (workbook_results.min_row + 1, workbook_results.max_row)):
        try:
            fc = ColdChainFacility()
            fc.code = row[0].value
            fc.name = row[4].value
            fc.district = row[1].value
            fc.sub_county = row[2].value
            ft = FacilityType.objects.get(old_id=row[3].value)
            fc.type = ft
            fc.save()
        except IntegrityError:
            print("%s already exists" % row[4].value)


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """ Import facility """

    def handle(self, *args, **options):
        excel_file = args[0]
        import_facilities(excel_file)