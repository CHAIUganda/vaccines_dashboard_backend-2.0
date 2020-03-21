from datetime import date
from django.core.management import BaseCommand
from django.db import IntegrityError

from cold_chain.models import *
from openpyxl import load_workbook


def import_eligible_facility_metrics(excel_file):
    """
    Extracts data from 'Eligible HF List by District + CCE Coverage_June 2019.xlsx'
    """
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True, data_only=True)
    # names with spaces some times cause error.
    # Replacing spaces with underscores prevents this error
    worksheet_name = "Eligible_HF_List_by_District"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('B%s:I%s' % (workbook_results.min_row + 1, workbook_results.max_row)):
        try:
            fc = EligibleFacilityMetric()
            fc.district = District.objects.filter(name__icontains=row[0].value).first()
            fc.total_eligible_facility = row[4].value
            fc.total_number_immunizing_facility = row[5].value
            fc.save()
        except Exception as e:
            print(e)


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """ Import facility """

    def handle(self, *args, **options):
        excel_file = args[0]
        import_facilities(excel_file)
