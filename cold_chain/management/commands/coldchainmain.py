from datetime import date
from django.core.management import BaseCommand
import traceback
from cold_chain.models import *
from openpyxl import load_workbook
import datetime
import random


def import_coldchainmain(excel_file, year, month):
    """
    Extracts data from ColdChainMain.xlsx
    """
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True, data_only=True)
    # # spaces in the sheet name may cause errors
    worksheet_name = "Functionality_and_Optimality"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:O%s' % (workbook_results.min_row + 1, workbook_results.max_row)):
        try:
            facility_code = row[1].value
            district = row[2].value
            facility_name = row[3].value
            type_of_facility = row[4].value
            total_population = row[5].value
            cce_model = row[6].value
            cce_make = row[7].value
            cce_serial_number = row[8].value
            available_net_storage_volume = row[9].value
            required_net_storage_volume = row[10].value
            supply_year = row[11].value
            older_than_ten_years = row[12].value
            suboptiomal_cce = row[13].value
            cce_functionality_status = row[14].value
            district = District.objects.filter(name__icontains=district).first()

            try:
                facility = ColdChainFacility()
                facility.name = facility_name
                facility.district = district
                facility.code = facility_code

                try:
                    facility_type = FacilityType.objects.filter(name=type_of_facility).first()
                except Exception as e:
                    print(e)
                    facility_type = FacilityType.objects.create(name=type_of_facility)
                facility.type = facility_type
                facility.save()
            except Exception as e:
                print(e)
                facility = ColdChainFacility.objects.get(code__icontains=facility_code)

            try:
                refrige = Refrigerator.objects.get(serial_number=cce_serial_number)
            except Exception as e:
                print(e)
                refrige = Refrigerator()
                refrige.make = cce_make
                refrige.model = cce_model
                refrige.serial_number = cce_serial_number if cce_serial_number != None else random.randint(26150562344, 29999999999)
                refrige.supply_year = datetime.datetime(supply_year, 1, 1)
                refrige.cold_chain_facility = facility
                refrige.save()

            refrige_detail = RefrigeratorDetail()
            refrige_detail.refrigerator = refrige
            refrige_detail.district = district
            refrige_detail.available_net_storage_volume = int(float(available_net_storage_volume)) if available_net_storage_volume != None else 0
            refrige_detail.required_net_storage_volume = int(float(required_net_storage_volume)) if required_net_storage_volume != None else 0
            refrige_detail.functionality_status = cce_functionality_status
            refrige_detail.year = year
            refrige_detail.month = month
            refrige_detail.save()
        except Exception as e:
            print(traceback.print_exc())
            print(e)

    worksheet_name = "Eligible_HF_List_by_District"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('B%s:I%s' % (workbook_results.min_row + 1, workbook_results.max_row)):
        try:
            fc = EligibleFacilityMetric()
            fc.district = District.objects.filter(name__icontains=row[0].value).first()
            fc.total_eligible_facility = int(row[4].value)
            fc.date = datetime.datetime(int(year), int(month), 1)
            fc.total_number_immunizing_facility = int(row[5].value)
            fc.save()
        except Exception as e:
            print(e)


class Command(BaseCommand):
    args = '<path to dataset file>, year, month'
    help = """ Import functionality. Imports the cold chain data 
            Pass the excel file link, year and period of the year"""

    def handle(self, *args, **options):
        import_coldchainmain(args[0], args[1], args[2])
