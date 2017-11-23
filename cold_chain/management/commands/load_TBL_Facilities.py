
from datetime import date
from django.core.management import BaseCommand

from cold_chain.models import *
from openpyxl import load_workbook


def import_tab_facilities(excel_file):
    # Extra parameters removed because String cells were returning None
    # Original: workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    workbook = load_workbook(excel_file)
    worksheet_name = "TBL_FACILITY_TYPE"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:J%s' % (workbook_results.min_row + 1, workbook_results.max_row)):
        name = row[1].value

        facility_type = FacilityType.objects.filter(name=name).last()
        if facility_type is None:
            facility_type = FacilityType()

        facility_type.old_id = row[0].value
        facility_type.name = name
        facility_type.group = row[7].value
        facility_type.save()


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """ Import facility_type """

    def handle(self, *args, **options):
        excel_file = args[0]
        import_tab_facilities(excel_file)