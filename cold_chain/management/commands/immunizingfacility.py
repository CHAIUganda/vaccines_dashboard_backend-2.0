from datetime import date
from django.core.management import BaseCommand

from cold_chain.models import *
from openpyxl import load_workbook


def import_immunizing_facilities(excel_file, quarter):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    worksheet_name = "tbl_facilities"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:M%s' % (workbook_results.min_row + 1, workbook_results.max_row)):
        try:
            fn = ImmunizingFacility()
            fn.static = row[10].value
            fn.outreach = row[11].value
            fn.ficc_storage = row[9].value
            fn.quarter = quarter
            fc = Facility.objects.get(code=row[0].value)
            fn.facility = fc
            fn.save()
        except FacilityType.DoesNotExist:
            fn.static = row[10].value
            fn.outreach = row[11].value
            fn.ficc_storage = row[9].value
            fn.quarter = quarter
            fn.save()


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """ Import ImmunizingFacility """

    def handle(self, *args, **options):
        excel_file = args[0]
        quarter = args[1]
        import_immunizing_facilities(excel_file, quarter)