from datetime import date
from django.core.management import BaseCommand
import traceback
from cold_chain.models import *
from openpyxl import load_workbook
import datetime
import random
from utility import add_months


def import_performance_management(excel_file, year, month):
    # data_only allow the return of values not formulae
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True, data_only=True)
    # spaces in the sheet name may cause errors
    worksheet_name = "Data"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:BF%s' % (workbook_results.min_row + 4, workbook_results.max_row)):
        try:
            immunization_component_name = row[1].value
            objective = row[2].value
            activity_name = row[3].value
            level = row[4].value
            funded = row[5].value
            funding_status = row[6].value
            verification = row[7].value
            priority = row[8].value if row[8].value else FUNDING_PRIORITY_LEVEL[0][0]
            organization_name = row[9].value
            activity_cost_ugx = row[10].value
            activity_cost_usd = row[11].value
            budget_assumption = row[12].value if row[12].value else ""
            responsible_focal_point = row[14].value
            stackholder_focal_point = row[15].value if row[15].value else ""

            try:
                organization = Organization.objects.get(name=organization_name)
            except Exception as e:
                print(e)
                organization = Organization(name=organization_name)
                organization.save()

            try:
                immunization_component = ImmunizationComponent.objects.get(name=immunization_component_name)
            except Exception as e:
                print(e)
                immunization_component = ImmunizationComponent(name=immunization_component_name)
                immunization_component.save()

            # precreate the dates in the ActivityDates table
            activity_dates = []
            activity_date = datetime.datetime(int(year), int(month), 1)

            for index, x in enumerate(range(16, 34)):
                if row[x].value:
                    print(activity_date)
                    saved_activity_dates = ActivityDates.objects.get(date=activity_date)
                    activity_dates.append(saved_activity_dates)
                activity_date = add_months(activity_date, 1)

            quarter = 1
            activity_statuses = []
            for index, x in enumerate(range(34, 46)):
                # skip odd number columns
                # because we pick one number and use it to pick the status and the comment
                if x % 2 != 0:
                    continue
                # reset the quarter when the year ends
                if x == 42:
                    quarter = 1
                status_value = row[x].value
                comment_value = row[x + 1].value
                activity_status = ActivityStatus(status=status_value, comment=comment_value, quarter=quarter)
                activity_status.save()
                activity_statuses.append(activity_status)
                quarter += 1

            activity = Activity(name=activity_name, funding_status=funding_status, activity_cost_ugx=activity_cost_ugx,
                                immunization_component=immunization_component, objective=objective, level=level,
                                funding_priority_level=priority, verification=verification,
                                activity_cost_usd=activity_cost_usd, budget_assumption=budget_assumption,
                                responsible_focal_point=responsible_focal_point,
                                stackholder_focal_point=stackholder_focal_point, organization=organization)
            activity.save()
            activity.activity_date.add(*activity_dates)
            activity.activity_status.add(*activity_statuses)
        except Exception as e:
            print(traceback.print_exc())
            print(e)


class Command(BaseCommand):
    args = '<path to dataset file>, year, month'
    help = """ Import performance management. Imports the performance management excel data. 
            Pass the excel file link, year and month of the year"""

    def handle(self, *args, **options):
        import_performance_management(args[0], args[1], args[2])
