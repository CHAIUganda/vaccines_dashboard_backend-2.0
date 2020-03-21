from datetime import date
from django.core.management import BaseCommand
import traceback
from cold_chain.models import *
from openpyxl import load_workbook
import datetime
import random


def import_functionality(excel_file, year, month):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    # spaces in the sheet name may cause errors
    worksheet_name = "Functionality_and_Optimality"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:O%s' % (workbook_results.min_row + 1, workbook_results.max_row)):
        try:
            facility_code = row[1].value
            district = row[2].value
            facility_name = row[3].value
            type_of_facility = row[4].value
            total_population = row[5].value
            cce_model = row[6].value
            cce_make = row[7].value
            cce_serial_number = row[8].value
            available_net_storage_volume = row[9].value
            required_net_storage_volume = row[10].value
            supply_year = row[11].value
            older_than_ten_years = row[12].value
            suboptiomal_cce = row[13].value
            cce_functionality_status = row[14].value

            district = District.objects.filter(name__icontains=district).first()

            try:
                facility = ColdChainFacility()
                facility.name = facility_name
                facility.district = district
                facility.code = facility_code

                facility_type = FacilityType.objects.filter(name__icontains=type_of_facility).first()
                facility.type = facility_type
                facility.save()
            except Exception as e:
                print(e)
                facility = ColdChainFacility.objects.get(code__icontains=facility_code)

            try:
                refrige = Refrigerator.objects.get(serial_number=cce_serial_number)
            except Exception as e:
                print(e)
                refrige = Refrigerator()
                refrige.make = cce_make
                refrige.model = cce_model
                refrige.serial_number = cce_serial_number
                refrige.supply_year = datetime.datetime(supply_year, 1, 1)
                refrige.cold_chain_facility = facility
                refrige.save()

            refrige_detail = RefrigeratorDetail()
            refrige_detail.refrigerator = refrige
            refrige_detail.district = district
            refrige_detail.temperature = random.randint(-4, 12)
            refrige_detail.available_net_storage_volume = available_net_storage_volume
            refrige_detail.required_net_storage_volume = required_net_storage_volume
            refrige_detail.functionality_status = cce_functionality_status
            refrige_detail.year = year
            refrige_detail.month = month
            refrige_detail.save()
        except Exception as e:
            print(traceback.print_exc())
            print(e)


class Command(BaseCommand):
    args = '<path to dataset file>, year, year_half'
    help = """ Import functionality. Imports the cold chain CCE data. 
            Pass the excel file link, year and period of the year"""

    def handle(self, *args, **options):
        import_functionality(args[0], args[1], args[2])
