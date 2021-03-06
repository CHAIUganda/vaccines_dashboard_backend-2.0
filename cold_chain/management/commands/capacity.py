from datetime import date
from django.core.management import BaseCommand
from django.db import IntegrityError

from cold_chain.models import *
from openpyxl import load_workbook


def import_capacity(excel_file, quarter):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    worksheet_name = "capacity"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:N%s' % (workbook_results.min_row + 4, workbook_results.max_row)):
        try:
            fc = Capacity()
            ft = ColdChainFacility.objects.get(code=row[0].value)
            fc.facility = ft
            fc.actual = row[8].value
            fc.required = row[9].value
            fc.difference = row[10].value
            fc.quarter = quarter
            fc.save()
        except ColdChainFacility.DoesNotExist:
            fc.actual = row[8].value
            fc.required = row[9].value
            fc.difference = row[10].value
            fc.quarter = quarter
            fc.save()
        except IntegrityError, e:
            print(e)


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """ Import capacity """

    def handle(self, *args, **options):
        excel_file = args[0]
        quarter = args[1]
        import_capacity(excel_file, quarter)