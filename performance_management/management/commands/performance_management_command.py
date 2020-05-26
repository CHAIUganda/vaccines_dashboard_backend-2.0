from datetime import date
from django.core.management import BaseCommand
import traceback
from performance_management.models import *
from openpyxl import load_workbook
import datetime
from utility import add_months, encode_unicode, quarter_months


def import_performance_management(excel_file, year, month):
    # data_only allow the return of values not formulae
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True, data_only=True)
    # spaces in the sheet name may cause errors
    # todo. replace with first sheet name
    worksheet_name = "Jan20-Jun21"
    workbook_results = workbook.get_sheet_by_name(worksheet_name)

    for row in workbook_results.iter_rows('A%s:BG%s' % (workbook_results.min_row + 4, workbook_results.max_row)):
        try:
            number = row[0].value
            immunization_component_name = row[1].value
            objective = row[2].value
            activity_name = row[3].value
            level = row[4].value
            funding_state = row[5].value
            funding_status = row[6].value
            verification = row[7].value
            priority = row[8].value if row[8].value else FUNDING_PRIORITY_LEVEL[0][0]
            organization_name = row[9].value
            funding_source_organization_name = row[10].value
            activity_cost_ugx = row[11].value
            activity_cost_usd = row[12].value
            budget_assumption = row[13].value if row[12].value else ""
            time_frame = row[14].value if row[12].value else ""
            responsible_focal_point = row[15].value if row[15].value else ""
            stackholder_focal_point = row[16].value if row[16].value else ""

            if not activity_name:
                continue

            try:
                organization = Organization.objects.get(name=organization_name.encode('ascii', 'ignore').decode('ascii'))
            except Exception as e:
                print(e)
                organization = Organization(name=organization_name)
                organization.save()

            try:
                funding_source_organization = FundingSourceOrganization.objects.get(name=funding_source_organization_name)
            except Exception as e:
                print(e)
                funding_source_organization = FundingSourceOrganization(name=funding_source_organization_name)
                funding_source_organization.save()

            try:
                immunization_component = ImmunizationComponent.objects.get(name=immunization_component_name)
            except Exception as e:
                print(e)
                immunization_component = ImmunizationComponent(name=immunization_component_name)
                immunization_component.save()

            activity_dates = []
            # activity date
            for index, x in enumerate(range(17, 35)):
                # incase the activity is active in the current month, add to list of activity_dates
                if row[x].value:
                    # calculates transition from one year to the next while resetting the month
                    activity_date = datetime.datetime(int(year) if index + 1 < 13 else int(year) + 1,
                                                      index + 1 if index + 1 < 13 else index + 1 - 12, 1)
                    # ActivityDates are shared by all activities
                    try:
                        saved_activity_dates = ActivityDates.objects.get(date=activity_date)
                    except ActivityDates.DoesNotExist as e:
                        print(e)
                        saved_activity_dates = ActivityDates(date=activity_date)
                        saved_activity_dates.save()
                    activity_dates.append(saved_activity_dates)

            # quarters
            active_quarters = []
            current_quarter = 1
            for index, x in enumerate(range(35, 41)):
                if row[x].value:
                    if current_quarter < 5:
                        active_quarters.append(current_quarter)
                    else:
                        current_quarter = 1
                        active_quarters.append(current_quarter)
                else:
                    active_quarters.append(0)
                current_quarter += 1

            # quarter budget
            budgets = []
            for x in range(41, 47):
                budgets.append(row[x].value if row[x].value > 2 else 0)

            quarter = 1
            activity_statuses = []
            # activity status
            counter = 0
            for index, x in enumerate(range(47, 58)):
                # skip odd number columns
                # because we pick one number and use it to pick the status and the comment
                # NOTE: addition of a single column will affect the quarter and skip logic
                if x % 2 == 0:
                    continue
                # reset the quarter when the year ends
                if x == 55:
                    quarter = 1

                if active_quarters[counter]:
                    status_value = row[x].value if row[x].value else COMPLETION_STATUS[2][0]
                    comment_value = row[x + 1].value if row[x + 1].value else ""
                    current_year = int(year) if index < 8 else int(year) + 1
                    current_month = quarter_months[active_quarters[counter]][0]
                    firstdate = datetime.datetime(current_year, current_month, 1)
                    activity_status = ActivityStatus(status=status_value, comment=comment_value, quarter=active_quarters[counter], year=current_year,
                                                     quarter_budget_usd=budgets[counter], firstdate=firstdate)
                    activity_status.save()
                    activity_statuses.append(activity_status)
                    quarter += 1
                counter += 1

            activity = Activity(name=activity_name, funding_status=funding_status, activity_cost_ugx=activity_cost_ugx,
                                immunization_component=immunization_component, objective=objective, level=level,
                                funding_priority_level=priority, verification=verification, funding_state=funding_state,
                                activity_cost_usd=activity_cost_usd, budget_assumption=budget_assumption,
                                funding_source_organization=funding_source_organization, time_frame=time_frame,
                                responsible_focal_point=responsible_focal_point, number=number,
                                stackholder_focal_point=stackholder_focal_point, organization=organization)
            activity.save()
            activity.activity_date.add(*activity_dates)
            activity.activity_status.add(*activity_statuses)
            print(vars(activity))
        except Exception as e:
            print(traceback.print_exc())
            print(e)


class Command(BaseCommand):
    args = '<path to dataset file>, year, month'
    help = """ Import performance management. Imports the performance management excel data. 
            Pass the excel file link, year and month of the year"""

    def handle(self, *args, **options):
        import_performance_management(args[0], args[1], args[2])
