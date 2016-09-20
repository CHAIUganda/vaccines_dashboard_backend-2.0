import logging
from collections import defaultdict
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
from dashboard.helpers import *
from dashboard.models import *

logger = logging.getLogger(__name__)


def get_real_facility_name(facility_name, district_name):
    if facility_name:
        template = "_%s" % district_name
        return facility_name.replace(template, "")
    else:
        return facility_name


class StockReport:
    def __init__(self, path, year, month):
        self.path = path
        self.year = year
        self.month = month
        self.month_name = MONTHS_TO_STR[int(self.month)]
        self.month_name = self.month_name.upper()
        self.vaccines = ["MEASLES", "BCG", "HPV", "BOPV", "TT", "TOPV", "IPV", "PCV", "PENTA"]
        self.period = int(year + month)

    def load(self):
        self.workbook = self.get_workbook()
        return self

    def get_workbook(self):
        return load_workbook(self.path, read_only=True, use_iterators=True)

    def import_balances(self):
        # Todo: use proper name
        worksheet_name = "%s %s" % (self.month_name, self.year)
        location_sheet = self.workbook.get_sheet_by_name(worksheet_name)
        current_period = self.period

        for row in location_sheet.iter_rows('B%s:K%s' % (location_sheet.min_row + 2, location_sheet.max_row)):
            if row[0].value:
                for vaccine in self.vaccines:
                    col = self.vaccines.index(vaccine) + 1
                    xls_row_number = row[0].row

                    if row[col].value:
                        if isFloat(row[col].value):
                            value = row[col].value
                        else:
                            value = float(row[col].value)
                    else:
                        value = 0

                    stock_requirement = StockRequirement.objects.filter(
                        district__name__contains=row[0].value,
                        vaccine__name=vaccine,
                        year=int(self.year),
                    ).first()

                    if stock_requirement:
                        if value:
                            try:
                                stock = Stock.objects.get(
                                    stock_requirement=stock_requirement,
                                    month=self.month)
                                stock.at_hand = value
                                stock.save()
                            except Stock.DoesNotExist:
                                stock = Stock()
                                stock.at_hand = value
                                stock.period = self.period
                                stock.firstdate = date(int(self.year), int(self.month), 1)
                                stock.lastdate = date(int(self.year), int(self.month), LAST_MONTH_DAY[int(self.month)])
                                stock.save()

    def import_orders(self):
        # Todo: use proper name
        worksheet_name = "%s %s" % (self.month_name, self.year)
        location_sheet = self.workbook.get_sheet_by_name(worksheet_name)
        for row in location_sheet.iter_rows('B%s:AC%s' % (location_sheet.min_row + 2, location_sheet.max_row)):
            if row[0].value:
                for vaccine in self.vaccines:
                    col = self.vaccines.index(vaccine) + 18
                    xls_row_number = row[0].row

                    if row[col].value:
                        if isFloat(row[col].value):
                            value = row[col].value
                        else:
                            value = float(row[col].value)
                    else:
                        value = 0

                    stock_requirement = StockRequirement.objects.filter(
                        district__name__contains=row[0].value,
                        vaccine__name=vaccine,
                        year=int(self.year),
                    ).first()

                    if stock_requirement:
                        if value:
                            try:
                                stock = Stock.objects.get(stock_requirement=stock_requirement, month=self.month)
                                stock.ordered = value
                                stock.save()
                            except Stock.DoesNotExist:
                                stock = Stock()
                                stock.ordered = value
                                stock.period = self.period
                                stock.firstdate = date(int(self.year), int(self.month), 1)
                                stock.lastdate = date(int(self.year), int(self.month), LAST_MONTH_DAY[int(self.month)])
                                stock.save()

    def import_issues(self):
        # Todo: use proper name
        worksheet_name = "ISSUED SUPPLY"
        location_sheet = self.workbook.get_sheet_by_name(worksheet_name)
        current_period = self.period

        for row in location_sheet.iter_rows('B%s:K%s' % (location_sheet.min_row + 1, location_sheet.max_row)):
            if row[0].value:
                for vaccine in self.vaccines:
                    col = self.vaccines.index(vaccine) + 1
                    xls_row_number = row[0].row

                    if row[col].value:
                        if isFloat(row[col].value):
                            value = row[col].value
                        else:
                            value = float(row[col].value)
                    else:
                        value = 0

                    stock_requirement = StockRequirement.objects.filter(
                        district__name__contains=row[0].value,
                        vaccine__name=vaccine,
                        year=int(self.year),
                    ).first()

                    if stock_requirement:
                        if value:
                            try:
                                stock = Stock.objects.get(
                                    stock_requirement=stock_requirement,
                                    month=self.month)
                                stock.received = value
                                stock.save()
                            except Stock.DoesNotExist:
                                stock = Stock()
                                stock.received = value
                                stock.period = self.period
                                stock.firstdate = date(int(self.year), int(self.month), 1)
                                stock.lastdate = date(int(self.year), int(self.month), LAST_MONTH_DAY[int(self.month)])
                                stock.save()

    def get_value(self, row, i):
        if i <= len(row):
            real_value = row[i].value
            value = real_value
            if value != "-" and value != '':
                return real_value
