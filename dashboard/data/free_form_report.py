import logging
from collections import defaultdict

from openpyxl import load_workbook

from dashboard.helpers import *
from dashboard.models import YearMonth

logger = logging.getLogger(__name__)


def get_real_facility_name(facility_name, district_name):
    if facility_name:
        template = "_%s" % district_name
        return facility_name.replace(template, "")
    else:
        return facility_name


class FreeFormReport():
    def __init__(self, path, year_month):
        self.path = path
        self.year_month = year_month
        self.name_cache = dict()
        self.districts = []

    def build_form_db(self, year_month):
        self.year_month = year_month.title
        return self

    def save(self):
        year_month, created = YearMonth.objects.get_or_create(title=self.year_month)
        year_month.save()
        return year_month

    def get_workbook(self):
        return load_workbook(self.path, read_only=True, use_iterators=True)

    def get_value(self, row, i):
        if i <= len(row):
            real_value = row[i].value
            value = real_value
            if value != "-" and value != '':
                return real_value

    def load(self):
        self.workbook = self.get_workbook()
        self.districts = self.districts()
        return self

    def districts(self):
        location_sheet = self.workbook.get_sheet_by_name(self.year_month.title)
        district_data = []
        for row in location_sheet.iter_rows('B%s:J%s' % (location_sheet.min_row + 3, location_sheet.max_row)):
            if row[0].value:
                district = dict()
                district[DISTRICT] = row[5].value
                district_data.append(district)
        return district_data

