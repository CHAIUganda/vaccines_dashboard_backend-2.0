# !/usr/bin/python
# -*- coding: utf-8 -*-

from django.core.management import BaseCommand

from dashboard.models import *
from django.db.models import Count, Sum, Avg
from openpyxl import load_workbook


def import_min_max(excel_file, year):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)

    vaccines = ["MEASLES", "BCG", "TT", "OPV",  "PCV", "PENTA", "HPV", "IPV"]

    for vaccine in vaccines:
        location_sheet = workbook.get_sheet_by_name(vaccine)
        for row in location_sheet.iter_rows('A6:G118'):
                district_object = District.objects.filter(name__contains=row[1].value).first()
                vaccine_object = Vaccine.objects.filter(name=vaccine).first()
                max_value = row[5].value
                #min_value = int(max_value) * .25
                min_value = row[6].value
                #print "district %s vaccine: %s min:%s max: %s" % (district_object.name, vaccine_object.name, min_value, max_value)
                if max_value:
                    stock_requirement = StockRequirement()
                    stock_requirement.district = district_object
                    stock_requirement.year = int(year)
                    stock_requirement.vaccine = vaccine_object
                    stock_requirement.minimum = min_value
                    stock_requirement.maximum = max_value
                    stock_requirement.save()


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """ Import stock requirements data  """

    def handle(self, *args, **options):
        excel_file = args[0]
        year = args[1]
        import_min_max(excel_file, year)
