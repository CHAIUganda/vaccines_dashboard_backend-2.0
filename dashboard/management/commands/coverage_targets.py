from datetime import date

from django.core.management import BaseCommand

from dashboard.models import *
from openpyxl import load_workbook
from dashboard.helpers import *


def import_coverage_target(excel_file, year):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    worksheet_name = "mcoverage"
    location_sheet = workbook.get_sheet_by_name(worksheet_name)

    # vaccines = ['BCG', 'OPV', 'PCV', 'TT', 'PENTA', 'MEASLES', 'IPV', 'HPV']
    vaccines = ['BCG', 'HPV', 'IPV', 'MEASLES', 'OPV', 'PCV', 'PENTA', 'TT', 'ROTA']

    for row in location_sheet.iter_rows('A%s:J%s' % (location_sheet.min_row + 5, location_sheet.max_row)):
        if row[0].value:

            for vaccine in vaccines:
                col = vaccines.index(vaccine) + 1

                if row[col].value:
                    if isFloat(row[col].value):
                        value = row[col].value
                    else:
                        value = float(row[col].value)
                else:
                    value = 0

                try:
                    stock_requirement = StockRequirement.objects.get(
                        district__name__contains=row[0].value,
                        vaccine__name=vaccine,
                        year=int(year))
                    stock_requirement.coverage_target = value
                    stock_requirement.save()
                    print 'coverage_target saved for %s - %s' % (row[0].value, vaccine)
                except StockRequirement.DoesNotExist:
                    district_object = District.objects.filter(name__contains=row[0].value).first()
                    vaccine_object = Vaccine.objects.filter(name=vaccine).first()
                    stock_requirement = StockRequirement()
                    stock_requirement.district = district_object
                    stock_requirement.year = int(year)
                    stock_requirement.vaccine = vaccine_object
                    stock_requirement.coverage_target = value
                    stock_requirement.save()
                    print 'coverage_target saved for %s - %s' % (row[0].value, vaccine)


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """Import stock requirements data. python manage.py coverage_target <file path> <year>"""

    def handle(self, *args, **options):
        excel_file = args[0]
        year = args[1]
        import_coverage_target(excel_file, year)
        print "Completed"
