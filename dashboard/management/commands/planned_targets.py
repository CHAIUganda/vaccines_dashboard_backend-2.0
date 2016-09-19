__author__ = 'tom'

from datetime import date

from django.core.management import BaseCommand

from dashboard.models import *
from openpyxl import load_workbook
from dashboard.helpers import *


def import_targets(excel_file, year):
    workbook = load_workbook(excel_file, read_only=True, use_iterators=True)
    worksheet_name = "Targets"
    location_sheet = workbook.get_sheet_by_name(worksheet_name)

    vaccines = ['BCG', 'PCV', 'TT', 'PENTA', 'MEASLES', 'IPV', 'HPV', 'BOPV', 'TOPV']

    for row in location_sheet.iter_rows('A%s:I%s' % (location_sheet.min_row + 5, location_sheet.max_row)):
        if row[0].value:
            for vaccine in vaccines:
                col = vaccines.index(vaccine) + 1

                if row[col].value:
                    if isFloat(row[col].value):
                        value = row[col].value
                    else:
                        value = float(row[col].value)
                else:
                    value = 0
                try:
                    stock_requirement = StockRequirement.objects.get(
                        district__name__contains=row[0].value,
                        vaccine__name=vaccine,
                        year=int(year))
                    stock_requirement.target = value
                    stock_requirement.save()
                    print 'Target saved for %s - %s' % (row[0].value, vaccine)
                except StockRequirement.DoesNotExist:
                    print 'No record found for %s - %s' % (row[0].value, vaccine)
                    pass


class Command(BaseCommand):
    args = '<path to dataset file>'
    help = """Import stock requirements data. python manage.py planned_targets <file path> <year>"""

    def handle(self, *args, **options):
        excel_file = args[0]
        year = args[1]
        import_targets(excel_file, year)
        print "Completed"
