from dashboard.models import *
from django.db.models import Count, Sum, Avg
from openpyxl import load_workbook


summary = Stock.objects.all().values('district__name').annotate(stockathand=Sum('at_hand')).order_by('district__name', 'period').values('district__name', 'stockathand', 'month', 'period')

excel_file = "c:\\Dev\\Python\\vaccines\\Docs\\Vaccines Dashboard Phase 1 Data v2\\Stock Management\\2015_12_30_CVS_Vax_Supplied.xlsx"
workbook = load_workbook(excel_file, read_only=True, use_iterators=True)

vaccines = ["MEASLES", "BCG", "HPV", "HEPB", "TT", "TOPV", "YELLOW FEVER", "PCV", "PENTA"]

for vaccine in self.vaccines:
    location_sheet = workbook.get_sheet_by_name(vaccine)
    for row in location_sheet.iter_rows('A6:F117'):
            district_object = District.objects.filter(name__contains=row[1].value).first()
            vaccine_object = Vaccine.objects.filter(name=vaccine).first()
            max_value = row[5].value
            min_value = int(max_value) * .25
            print "district %s vaccine: %s min:%s max: %s" % (district_object.name, vaccine_object.name, min_value, max_value)
            stock_requirement = StockRequirement()
            stock_requirement.district = district_object
            stock_requirement.year = 2015
            stock_requirement.vaccine = vaccine_object
            stock_requirement.minimum = min_value
            stock_requirement.maximum = max_value
            stock_requirement.save()

